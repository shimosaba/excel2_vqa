"""
STEP 1: Excel → シート別 PNG 画像化
LibreOffice でPDF変換 → pdf2image でページ別PNG化
"""
from __future__ import annotations

import subprocess
import tempfile
from pathlib import Path

import openpyxl
from pdf2image import convert_from_path
from PIL import Image, ImageChops

from .config import Config


def _calc_dpi(ws) -> int:
    """シートのセル範囲からDPIを動的計算する。大きいシートほど高DPI。"""
    cells = (ws.max_row or 1) * (ws.max_column or 1)
    if cells < 50:
        return 300
    elif cells < 300:
        return 350
    elif cells < 1000:
        return 400
    elif cells < 3000:
        return 450
    else:
        return 500


def _trim_whitespace(img: Image.Image, padding: int = 10) -> Image.Image:
    """白い余白を除去してデータ領域のみにクロップする。"""
    gray = img.convert("L")
    diff = ImageChops.difference(gray, Image.new("L", gray.size, 255))
    bbox = diff.getbbox()
    if bbox is None:
        return img  # 全白の場合はそのまま返す
    pw, ph = img.size
    x1 = max(0, bbox[0] - padding)
    y1 = max(0, bbox[1] - padding)
    x2 = min(pw, bbox[2] + padding)
    y2 = min(ph, bbox[3] + padding)
    return img.crop((x1, y1, x2, y2))


def render_workbook(excel_path: Path, output_dir: Path, config: Config) -> list[Path]:
    """
    Excel を1シートずつ PNG に変換して返す。

    Args:
        excel_path: 入力Excelファイルパス
        output_dir: 出力先ディレクトリ
        config: 設定オブジェクト

    Returns:
        生成したPNGファイルパスのリスト（シート順）
    """
    excel_path = Path(excel_path).resolve()
    sheets_dir = output_dir / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n[STEP 1] Excel → シート画像化: {excel_path.name}")

    # openpyxl でシート名を取得
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    with tempfile.TemporaryDirectory() as tmpdir:
        # LibreOffice で PDF 変換
        result = subprocess.run(
            [
                "libreoffice", "--headless",
                "--convert-to", "pdf",
                "--outdir", tmpdir,
                str(excel_path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice 変換失敗:\n{result.stderr}")

        pdf_files = list(Path(tmpdir).glob("*.pdf"))
        if not pdf_files:
            raise RuntimeError("PDF が生成されませんでした")
        pdf_path = pdf_files[0]

        # シートごとに DPI 計算 → PNG 変換 → 余白除去
        image_paths: list[Path] = []
        wb2 = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        for i, sname in enumerate(sheet_names):
            ws = wb2[sname]
            sheet_dpi = config.dpi if config.dpi > 0 else _calc_dpi(ws)
            img_path = sheets_dir / f"sheet_{i + 1:02d}.png"

            pages = convert_from_path(str(pdf_path), dpi=sheet_dpi,
                                      first_page=i + 1, last_page=i + 1)
            orig = pages[0].convert("RGB")
            page_rgb = _trim_whitespace(orig)
            page_rgb.save(str(img_path), "PNG")
            image_paths.append(img_path)

            print(f"  → sheet_{i + 1:02d}.png  [{sname}]  "
                  f"DPI={sheet_dpi}  {orig.width}x{orig.height}px → "
                  f"{page_rgb.width}x{page_rgb.height}px (余白除去後)")
        wb2.close()

    print(f"  合計 {len(image_paths)} シート画像を生成しました")
    return image_paths
